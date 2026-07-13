import json, sys, os
HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.dirname(os.path.dirname(HERE))
sys.path.insert(0, HERE)
from i18n_categories import sub_for_key
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

recs = json.load(open(os.path.join(HERE, 'records.json'), encoding='utf-8'))

dict_recs = [r for r in recs if r['kind'] == 'قاموس I18N']
inline_recs = [r for r in recs if r['kind'] != 'قاموس I18N']

# ---- attach sub-location for dict recs ----
for r in dict_recs:
    sub_ar, sub_en = sub_for_key(r['key'])
    r['main_ar'] = 'قاموس الواجهة (يظهر في كل الصفحات)'
    r['main_en'] = 'Interface dictionary (used across all pages)'
    r['ctx'] = sub_ar
    r['ctx_en'] = sub_en

for r in inline_recs:
    r['ctx_en'] = r.get('ctx') or ''

# ---- sort ----
# build ordered list of main sections as they appear (by first-seen line) for inline sheet
main_order = []
seen_main = set()
for r in sorted(inline_recs, key=lambda x: x['line']):
    if r['main_ar'] not in seen_main:
        seen_main.add(r['main_ar'])
        main_order.append(r['main_ar'])
main_rank = {m: i for i, m in enumerate(main_order)}
inline_recs.sort(key=lambda r: (main_rank[r['main_ar']], r['line']))

dict_order = [c[0] for c in __import__('i18n_categories').CATS]
dict_rank = {c: i for i, c in enumerate(dict_order)}
dict_recs.sort(key=lambda r: (dict_rank.get(r['ctx'], 999), r['key']))

# ================= STYLES =================
NAVY = "1F3864"
GOLD = "B8860B"
LIGHT_GOLD = "FDF3DC"
LIGHT_NAVY = "EAF0F8"
WHITE = "FFFFFF"
GRAY_BORDER = "C9C9C9"

title_font = Font(name="Calibri", size=16, bold=True, color=WHITE)
subtitle_font = Font(name="Calibri", size=11, italic=True, color=WHITE)
header_font = Font(name="Calibri", size=11, bold=True, color=WHITE)
main_section_font = Font(name="Calibri", size=12, bold=True, color=WHITE)
sub_font = Font(name="Calibri", size=10, bold=True, color="1F3864")
body_font = Font(name="Calibri", size=11, color="222222")
edit_font = Font(name="Calibri", size=11, color="0B5A2B", bold=True)
note_font = Font(name="Calibri", size=10, italic=True, color="666666")

header_fill = PatternFill("solid", fgColor=NAVY)
main_fill = PatternFill("solid", fgColor=NAVY)
sub_fill = PatternFill("solid", fgColor=LIGHT_GOLD)
edit_fill = PatternFill("solid", fgColor="EAF7EC")
alt_fill = PatternFill("solid", fgColor=LIGHT_NAVY)
title_fill = PatternFill("solid", fgColor=NAVY)

thin = Side(style="thin", color=GRAY_BORDER)
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_header_row(ws, row, ncols, fill=header_fill, font=header_font, height=26):
    ws.row_dimensions[row].height = height
    for c in range(1, ncols+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_all

def autofit(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

wb = Workbook()

# ================= SHEET 0: دليل الاستخدام (Guide / TOC) =================
ws0 = wb.active
ws0.title = "دليل الاستخدام"
ws0.sheet_view.rightToLeft = True
ws0.merge_cells("A1:F1")
ws0["A1"] = "ملف المسميّات الرئيسي — منصة السجل الوطني للتراث العمراني"
ws0["A1"].font = title_font
ws0["A1"].fill = title_fill
ws0["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws0.row_dimensions[1].height = 34

ws0.merge_cells("A2:F2")
ws0["A2"] = "الإصدار 3 — بعد حذف ميزات الاستوديو التسع (أُبقي: الاستيراد الذكي + آلة الزمن) · باستثناء نصوص «المساعد الذكي» بالكامل"
ws0["A2"].font = subtitle_font
ws0["A2"].fill = title_fill
ws0["A2"].alignment = Alignment(horizontal="center", vertical="center")
ws0.row_dimensions[2].height = 22

guide_lines = [
    ("", ""),
    ("كيف تستخدم هذا الملف", ""),
    ("1. الورقة «2- قاموس الواجهة» تضم كل النصوص الثابتة العامة (أزرار، عناوين، رسائل) المستخدمة في أكثر من مكان بالموقع.", ""),
    ("2. الورقة «3- النصوص المضمّنة» تضم كل النصوص الخاصة بميزة أو شاشة معيّنة (مرتّبة حسب موقعها الفعلي في الموقع).", ""),
    ("3. في كل ورقة: عمود «القسم الرئيسي» = الشاشة/الميزة، وعمود «الموقع الفرعي» = التفصيل داخلها (مثال: زر، حقل، رسالة تنبيه).", ""),
    ("4. عدّل فقط عمودي «الجديد (عربي)» و«الجديد (English)» — اترك «الحالي» كما هو تماماً (يُستخدم للمطابقة عند الاستيراد).", ""),
    ("5. إن لم ترغب بتغيير نص معيّن، اترك خانة «الجديد» فارغة — سيبقى كما هو.", ""),
    ("6. الصفوف مجمّعة (Group) حسب القسم الرئيسي — اضغط على [-] يسار الصفوف لطي كل قسم وتصغيره، أو استخدم أزرار المستويات ١/٢ أعلى يسار الشاشة.", ""),
    ("7. النصوص المعلّمة «ديناميكي» تحتوي على قيم متغيّرة مثل ${العدد} — لا تحذف هذا الجزء عند الترجمة، فقط عدّل النص المحيط به.", ""),
    ("8. بعد التعديل، أرسل نفس هذا الملف وسيتم عكس كل التغييرات على الموقع تلقائياً بمطابقة عمود «الحالي».", ""),
    ("", ""),
    ("ملاحظة هامة", ""),
    ("تم استثناء كل نصوص «المساعد الذكي» (Smart Assistant) من هذا الملف بناءً على طلبك — لن تجد أي نص متعلق به هنا.", ""),
]
r = 4
for line, _ in guide_lines:
    if line and not line[0].isdigit() and line not in ("ملاحظة هامة","كيف تستخدم هذا الملف"):
        ws0.cell(row=r, column=1, value=line).font = note_font
    elif line in ("ملاحظة هامة","كيف تستخدم هذا الملف"):
        ws0.cell(row=r, column=1, value=line).font = Font(size=13, bold=True, color=NAVY)
    else:
        ws0.cell(row=r, column=1, value=line).font = body_font
    ws0.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    ws0.cell(row=r, column=1).alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
    r += 1

r += 1
ws0.cell(row=r, column=1, value="فهرس الأقسام الرئيسية").font = Font(size=13, bold=True, color=NAVY)
ws0.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
r += 1
toc_header_row = r
headers0 = ["#", "القسم الرئيسي", "Main Section", "عدد النصوص", "الورقة"]
for i, h in enumerate(headers0, start=1):
    ws0.cell(row=r, column=i, value=h)
style_header_row(ws0, toc_header_row, 5, height=22)
r += 1

from collections import Counter, OrderedDict
dict_count = len(dict_recs)
sec_counts = OrderedDict()
sec_counts["قاموس الواجهة (يظهر في كل الصفحات)"] = ("Interface dictionary", dict_count, "2- قاموس الواجهة")
inline_counter = Counter(r_['main_ar'] for r_ in inline_recs)
for m in main_order:
    en = next(r_['main_en'] for r_ in inline_recs if r_['main_ar'] == m)
    sec_counts[m] = (en, inline_counter[m], "3- النصوص المضمّنة")

idx = 1
toc_start = r
for name_ar, (name_en, cnt, sheet) in sec_counts.items():
    ws0.cell(row=r, column=1, value=idx)
    ws0.cell(row=r, column=2, value=name_ar)
    ws0.cell(row=r, column=3, value=name_en)
    ws0.cell(row=r, column=4, value=cnt)
    ws0.cell(row=r, column=5, value=sheet)
    fill = alt_fill if idx % 2 == 0 else PatternFill("solid", fgColor=WHITE)
    for c in range(1, 6):
        cell = ws0.cell(row=r, column=c)
        cell.fill = fill
        cell.font = body_font
        cell.border = border_all
        cell.alignment = Alignment(horizontal="center" if c in (1,4) else "right", vertical="center", wrap_text=True)
    idx += 1
    r += 1

total_row = r
ws0.cell(row=r, column=1, value="")
ws0.cell(row=r, column=2, value="الإجمالي")
ws0.cell(row=r, column=4, value=len(recs))
for c in range(1, 6):
    cell = ws0.cell(row=r, column=c)
    cell.font = Font(bold=True, color=NAVY)
    cell.fill = PatternFill("solid", fgColor=LIGHT_GOLD)
    cell.border = border_all

autofit(ws0, [5, 42, 34, 12, 20])
ws0.freeze_panes = None

# ================= SHEET 1: قاموس الواجهة =================
ws1 = wb.create_sheet("2- قاموس الواجهة")
ws1.sheet_view.rightToLeft = True
headers1 = ["#", "الموقع الفرعي (فئة الاستخدام)", "Sub-location (usage)", "المفتاح البرمجي", "الحالي (عربي)", "الحالي (English)", "الجديد (عربي)", "الجديد (English)"]
for i, h in enumerate(headers1, start=1):
    ws1.cell(row=1, column=i, value=h)
style_header_row(ws1, 1, len(headers1))
ws1.freeze_panes = "A2"

row = 2
current_sub = None
group_start = None
n = 1
for r_ in dict_recs:
    if r_['ctx'] != current_sub:
        # sub-header row
        ws1.cell(row=row, column=1, value="")
        ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers1))
        cell = ws1.cell(row=row, column=1, value=f"▸ {r_['ctx']}  /  {r_['ctx_en']}")
        cell.font = sub_font
        cell.fill = sub_fill
        cell.alignment = Alignment(horizontal="right", vertical="center")
        ws1.row_dimensions[row].height = 20
        current_sub = r_['ctx']
        row += 1
    vals = [n, r_['ctx'], r_['ctx_en'], r_['key'], r_['ar'], r_['en'], "", ""]
    for c, v in enumerate(vals, start=1):
        cell = ws1.cell(row=row, column=c, value=v)
        cell.border = border_all
        cell.alignment = Alignment(horizontal="right" if c in (2,3,5,6,7,8) else "center", vertical="center", wrap_text=True)
        if c in (7, 8):
            cell.font = edit_font
            cell.fill = edit_fill
        else:
            cell.font = body_font
            cell.fill = PatternFill("solid", fgColor=WHITE) if n % 2 else alt_fill
    n += 1
    row += 1

autofit(ws1, [5, 26, 26, 18, 40, 40, 40, 40])
ws1.column_dimensions['A'].width = 5

# ================= SHEET 2: النصوص المضمّنة =================
ws2 = wb.create_sheet("3- النصوص المضمّنة")
ws2.sheet_view.rightToLeft = True
headers2 = ["#", "القسم الرئيسي", "الموقع الفرعي (الدالة/الميزة)", "نوع النص", "رقم السطر", "الحالي (عربي)", "الحالي (English)", "الجديد (عربي)", "الجديد (English)"]
for i, h in enumerate(headers2, start=1):
    ws2.cell(row=1, column=i, value=h)
style_header_row(ws2, 1, len(headers2))
ws2.freeze_panes = "A2"

row = 2
current_main = None
n = 1
outline_level = 0
for r_ in inline_recs:
    if r_['main_ar'] != current_main:
        ws2.cell(row=row, column=1, value="")
        ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers2))
        cell = ws2.cell(row=row, column=1, value=f"■ {r_['main_ar']}  /  {r_['main_en']}")
        cell.font = main_section_font
        cell.fill = main_fill
        cell.alignment = Alignment(horizontal="right", vertical="center")
        ws2.row_dimensions[row].height = 22
        current_main = r_['main_ar']
        row += 1
    ctx_display = r_['ctx'] or "(مستوى عام بالشاشة)"
    vals = [n, r_['main_ar'], ctx_display, r_['kind'].replace('قاموس I18N',''), r_['line'], r_['ar'], r_['en'], "", ""]
    for c, v in enumerate(vals, start=1):
        cell = ws2.cell(row=row, column=c, value=v)
        cell.border = border_all
        cell.alignment = Alignment(horizontal="right" if c in (2,3,4,6,7,8,9) else "center", vertical="center", wrap_text=True)
        if c in (8, 9):
            cell.font = edit_font
            cell.fill = edit_fill
        else:
            cell.font = body_font
            cell.fill = PatternFill("solid", fgColor=WHITE) if n % 2 else alt_fill
    ws2.row_dimensions[row].outline_level = 1
    n += 1
    row += 1

ws2.sheet_properties.outlinePr.summaryBelow = False
autofit(ws2, [5, 32, 26, 22, 10, 42, 42, 42, 42])
ws2.column_dimensions['A'].width = 5

# sheet order
wb._sheets = [ws0, ws1, ws2]

out_path = os.path.join(REPO, "TRANSLATIONS_MASTER.xlsx")
wb.save(out_path)
print("saved", out_path)
print("dict rows:", len(dict_recs), "inline rows:", len(inline_recs), "total:", len(dict_recs)+len(inline_recs))
